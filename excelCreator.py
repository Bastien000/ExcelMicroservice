def get_usage_by_id(usage_id):
    import requests

    url = f"http://localhost:8080/getUsage/{usage_id}"
    response = requests.get(url)
    response.raise_for_status()
    return response.json()

def get_building_by_id(building_id):
    import requests

    url = f"http://localhost:8080/getBuilding/{building_id}"
    response = requests.get(url)
    response.raise_for_status()
    return response.json()



def createExcel(json_data):
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import json
    import urllib.request
    contents = urllib.request.urlopen("http://192.168.1.137:8080/getCreatorOption/1").read()
    jsonOptions = contents.decode("utf-8")
    parsed = json.loads(jsonOptions)
    creatorOptions = parsed.get("creatorOptions", {})

    class TenantEntity:
        def __init__(self, tenant_entity_id, name, move_in_date, move_out_date, building_entity_id, floor, area, usage_entity_id):
            self.tenant_entity_id = tenant_entity_id
            self.name = name
            self.move_in_date = datetime.fromtimestamp(move_in_date / 1000) if isinstance(move_in_date, (int, float)) else move_in_date
            self.move_out_date = move_out_date
            self.building_entity_id = building_entity_id
            self.floor = floor
            self.area = area
            self.usage_entity_id = usage_entity_id

    tenants = []
    buildingsID = []
    for tenant_data in json_data:
        tenant_dict, usage_id = tenant_data
        tenant = TenantEntity(
            tenant_dict["tenantEntityID"],
            tenant_dict["name"],
            tenant_dict["moveInDate"],
            tenant_dict["moveOutDate"],
            tenant_dict["buildingEntityID"],
            tenant_dict["floor"],
            tenant_dict["area"],
            tenant_dict["usageEntityID"]
        )
        if tenant.building_entity_id not in buildingsID:
            buildingsID.append(tenant.building_entity_id)
        tenants.append(tenant)
    
    buildings = [get_building_by_id(id) for id in buildingsID]

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    if not buildings:
        workbook.create_sheet("default")

    building_colors = [PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
                       PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")]
    tenant_colors = [PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"),
                     PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")]

    for build in buildings:
        sheet = workbook.create_sheet(title=build.get("name", ""))
        headers = ["Firma", "Spotřeba", "Typ", "Rozloha", "Procentuální rozloha", "Fakturovat"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        row = 2
        sheet.cell(row=row, column=1).value = "Celkem"
        sheet.cell(row=row, column=2).value = f"=SUM(B{row+1}:B1048576)"
        sheet.cell(row=row, column=3).value = "None"
        sheet.cell(row=row, column=4).value = build.get("area", "")
        sheet.cell(row=row, column=5).value = "100%"
        row += 1

        tenants.sort(key=lambda t: t.floor)
        for tenant in tenants:
            if build.get("buildingEntityID", "") == tenant.building_entity_id:
                usage_info = get_usage_by_id(tenant.usage_entity_id[0])
                sheet.cell(row=row, column=1).value = tenant.name
                sheet.cell(row=row, column=2).value = usage_info.get("value", "")
                sheet.cell(row=row, column=3).value = usage_info.get("type", "")
                sheet.cell(row=row, column=4).value = tenant.area
                sheet.cell(row=row, column=5).value = f'=ROUND({tenant.area}/{build.get("area", 1)}*100, 2)&"%"'
                row += 1

        if build.get("name", "") in creatorOptions.get("buildings", {}):
            building_options = creatorOptions["buildings"][build.get("name", "")]
            if "add_rows" in building_options:
                for add_row in building_options["add_rows"]:
                    sheet.cell(row=row, column=1).value = add_row["name"]
                    sheet.cell(row=row, column=2).value = add_row["consumption"]
                    sheet.cell(row=row, column=3).value = add_row["type"]
                    sheet.cell(row=row, column=4).value = add_row["area"]
                    sheet.cell(row=row, column=5).value = add_row["percentage"]
                    row += 1

            if "apply_formulas" in building_options:
                for rule in building_options["apply_formulas"]:
                    condition = rule["condition"]
                    condition_row = None
                    for r in range(2, row):
                        cell_value = sheet.cell(row=r, column=1).value
                        if cell_value and condition["row_name_contains"].upper() in str(cell_value).upper():
                            condition_row = r
                            break
                    
                    if condition_row:
                        if "target_rows" in rule:
                            target_rows_spec = rule["target_rows"]
                            for r in range(2, row):
                                cell_value = sheet.cell(row=r, column=1).value
                                if "exclude" in target_rows_spec and cell_value in target_rows_spec["exclude"]:
                                    continue
                                if "include" in target_rows_spec and cell_value not in target_rows_spec["include"]:
                                    continue
                                formula = rule["formula"]
                                if isinstance(formula, dict) and cell_value in formula:
                                    current_value = sheet.cell(row=r, column=6).value or 0
                                    condition_value = sheet.cell(row=condition_row, column=2).value or 0
                                    sheet.cell(row=r, column=6).value = formula[cell_value].format(current_value=current_value, condition_value=condition_value)
                                else:
                                    sheet.cell(row=r, column=6).value = formula.format(row=r, condition_row=condition_row)
                        
                        if "action" in rule and rule["action"]["type"] == "modify_row":
                            action = rule["action"]
                            new_name = action["new_name"]
                            depends_on = rule.get("depends_on", {})
                            if "row_name_contains" in depends_on:
                                dep_row = None
                                for r in range(2, row):
                                    if depends_on["row_name_contains"].upper() in str(sheet.cell(row=r, column=1).value or "").upper():
                                        dep_row = r
                                        break
                                if dep_row:
                                    sheet.cell(row=condition_row, column=1).value = new_name
                                    if "new_fakturovat" in action:
                                        sheet.cell(row=condition_row, column=6).value = action["new_fakturovat"].format(row=condition_row, vp6_row=dep_row)

        for r in range(3, row):
            fill = tenant_colors[0] if r % 2 == 0 else tenant_colors[1]
            for c in range(1, 7):
                sheet.cell(row=r, column=c).fill = fill

    summary_sheet = workbook.create_sheet("Souhrn")
    summary_sheet.cell(row=1, column=1).value = "Budova"
    summary_sheet.cell(row=1, column=2).value = "Firma"
    summary_sheet.cell(row=1, column=3).value = "Fakturovat"
    for col in range(1, 4):
        summary_sheet.cell(row=1, column=col).font = Font(bold=True)

    summary_row = 2
    color_index = 0
    for build in buildings:
        building_name = build.get("name", "")
        building_sheet = workbook[building_name]
        building_color = building_colors[color_index % len(building_colors)]
        color_index += 1

        r = 3
        while building_sheet.cell(row=r, column=1).value:
            summary_sheet.cell(row=summary_row, column=1).value = building_name
            summary_sheet.cell(row=summary_row, column=2).value = f"='{building_name}'!A{r}"
            summary_sheet.cell(row=summary_row, column=3).value = f"='{building_name}'!F{r}"
            for c in range(1, 4):
                summary_sheet.cell(row=summary_row, column=c).fill = building_color
            summary_row += 1
            r += 1

    workbook.save("temp_excel_downloads/tenant_data.xlsx")
    print("Excel soubor 'tenant_data.xlsx' byl vytvořen.")